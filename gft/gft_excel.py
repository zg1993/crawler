# -*-codind: utf-8 -*-
'''
赣服通一件事非链办事项处理流程：
1、读取事项编码清单清单生成itemCodeRegionMap
2、找lym同步所有事项编码（可能需要生成应用名称、区县、事项编码excel文档）
3、调用表单提交接口确认问题事项编码，找一窗确认
'''

from openpyxl import load_workbook


def read2_20():
    workbook = load_workbook(
        '/home/zg/Downloads/work/oneThing/2-20一件事事项及经办人员(1).xlsx')
    sheet1 = workbook.worksheets[0]
    count = 0

    for row in range(2, 25):
        code = sheet1.cell(row, 5).value
        if (code):
            count = count + 1
            print('"{}",'.format(code))
    print(count)


def read_excel():
    workbook = load_workbook(
        '/home/zg/Downloads/work/oneThing/中考优待事项编码清单-215.xlsx')
    sheet1 = workbook.worksheets[0]
    count = 0
    for row in range(3, 33):
        town = sheet1.cell(row, 3).value
        if (town):
            count = count + 1
            code = sheet1.cell(row, 4).value
            print(town, ":", '"' + code + '",')
    print('count: ', count)


if __name__ == '__main__':
    read2_20()
    # read_excel()
