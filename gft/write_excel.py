# -*-codind: utf-8 -*-

from openpyxl import load_workbook

from jyjz_config import EMPYT

from datetime import datetime

date_today = datetime.now().strftime('%m-%d')


def write_jyjz_excel(error):
    workbook = load_workbook(
        '/home/zg/Downloads/work/oneThing/jyjz/教育救助问题乡镇.xlsx')
    sheet1 = workbook['Sheet1']
    for index, item in enumerate(EMPYT, 2):
        name, code = item
        sheet1.cell(index, 1).value = name
        sheet1.cell(index, 2).value = str(code)
        sheet1.cell(index, 3).value = error
    workbook.save(
        '/home/zg/Downloads/work/oneThing/jyjz/教育救助问题乡镇_{}.xlsx'.format(
            date_today))


if __name__ == '__main__':
    error = 'ItemId参数不能为空或未查询到ItemCode所对应的事项!'
    write_jyjz_excel(error)
