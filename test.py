# -*- coding: utf-8 -*-

# encoding_test.py

import sys
import locale

print(sys.version)
print(sys.executable)

print("=== 编码测试 ===")
print("文件编码声明:", __doc__)
print("系统默认编码:", sys.getdefaultencoding())
print("标准输出编码:", sys.stdout.encoding)
print("区域设置:", locale.getpreferredencoding())
print("数字测试:", 123)
print("英文测试:", "Hello")
print("中文测试:", "你好")
print("特殊符号测试:", "★")

# # 测试文件读写
# with open("test.txt", "w", encoding="utf-8") as f:
#     f.write("Python文件读写测试")

# with open("test.txt", "r", encoding="utf-8") as f:
#     print("文件内容:", f.read())
TEXT = '''
id_court_bad	string	通过身份证查询，法院失信人
id_court_executed	string	通过身份证查询，法院被执行人
id_bank_bad	string	通过身份证查询，银行(含信用卡)中风险
id_bank_overdue	string	通过身份证查询，银行(含信用卡)一般风险
id_bank_lost	string	通过身份证查询，银行(含信用卡)高风险
id_nbank_bad	string	通过身份证查询，非银(含全部非银类型)中风险
id_nbank_overdue	string	通过身份证查询，非银(含全部非银类型)一般风险
id_nbank_lost	string	通过身份证查询，非银(含全部非银类型)高风险
id_nbank_nsloan_bad	string	通过身份证查询，非银-持牌网络小贷中风险
id_nbank_nsloan_overdue	string	通过身份证查询，非银-持牌网络小贷一般风险
id_nbank_nsloan_lost	string	通过身份证查询，非银-持牌网络小贷高风险
id_nbank_sloan_bad	string	通过身份证查询，非银-持牌小贷中风险
id_nbank_sloan_overdue	string	通过身份证查询，非银-持牌小贷一般风险
id_nbank_sloan_lost	string	通过身份证查询，非银-持牌小贷高风险
id_nbank_cons_bad	string	通过身份证查询，非银-持牌消费金融中风险
id_nbank_cons_overdue	string	通过身份证查询，非银-持牌消费金融一般风险
id_nbank_cons_lost	string	通过身份证查询，非银-持牌消费金融高风险
id_nbank_finlea_bad	string	通过身份证查询，非银-持牌融资租赁中风险
id_nbank_finlea_overdue	string	通过身份证查询，非银-持牌融资租赁一般风险
id_nbank_finlea_lost	string	通过身份证查询，非银-持牌融资租赁高风险
id_nbank_autofin_bad	string	通过身份证查询，非银-持牌汽车金融中风险
id_nbank_autofin_overdue	string	通过身份证查询，非银-持牌汽车金融一般风险
id_nbank_autofin_lost	string	通过身份证查询，非银-持牌汽车金融高风险
id_nbank_other_bad	string	通过身份证查询，非银-其他中风险
id_nbank_other_overdue	string	通过身份证查询，非银-其他一般风险
id_nbank_other_lost	string	通过身份证查询，非银-其他高风险
'''
TEXT1 = '''
Rule_name_odr0000331	string	近两年命中法院失信人
Rule_weight_odr0000331	string	近两年命中法院失信人权重
Rule_name_odr0000332	string	两年前命中法院失信人
Rule_weight_odr0000332	string	两年前命中法院失信人权重
Rule_name_odr0000333	string	近两年命中法院被执行人
Rule_weight_odr0000333	string	近两年命中法院被执行人权重
Rule_name_odr0000334	string	两年前命中法院被执行人
Rule_weight_odr0000334	string	两年前命中法院被执行人权重
Rule_name_odr0000335	string	近两年命中银行高风险
Rule_weight_odr0000335	string	近两年命中银行高风险权重
Rule_name_odr0000336	string	两年前命中银行高风险
Rule_weight_odr0000336	string	两年前命中银行高风险权重
Rule_name_odr0000337	string	近两年命中银行中风险
Rule_weight_odr0000337	string	近两年命中银行中风险权重
Rule_name_odr0000338	string	两年前命中银行中风险
Rule_weight_odr0000338	string	两年前命中银行中风险权重
Rule_name_odr0000339	string	近两年命中银行一般风险
Rule_weight_odr0000339	string	近两年命中银行一般风险权重
Rule_name_odr0000340	string	两年前命中银行一般风险
Rule_weight_odr0000340	string	两年前命中银行一般风险权重
Rule_name_odr0000341	string	近两年命中非银高风险
Rule_weight_odr0000341	string	近两年命中非银高风险权重
Rule_name_odr0000342	string	两年前命中非银高风险
Rule_weight_odr0000342	string	两年前命中非银高风险权重
Rule_name_odr0000343	string	近两年命中非银中风险
Rule_weight_odr0000343	string	近两年命中非银中风险权重
Rule_name_odr0000344	string	两年前命中非银中风险
Rule_weight_odr0000344	string	两年前命中非银中风险权重
Rule_name_odr0000345	string	近两年命中非银一般风险
Rule_weight_odr0000345	string	近两年命中非银一般风险权重
Rule_name_odr0000346	string	两年前命中非银一般风险
Rule_weight_odr0000346	string	两年前命中非银一般风险权重
Rule_name_odr0000347	string	命中银行中风险次数较多
Rule_weight_odr0000347	string	命中银行中风险次数较多权重
Rule_name_odr0000348	string	命中银行一般风险次数较多
Rule_weight_odr0000348	string	命中银行一般风险次数较多权重
Rule_name_odr0000349	string	命中非银中风险次数较多
Rule_weight_odr0000349	string	命中非银中风险次数较多权重
Rule_name_odr0000350	string	命中非银一般风险次数较多
Rule_weight_odr0000350	string	命中非银一般风险次数较多权重

'''


from openpyxl import load_workbook



def read_excel():
    workbook = load_workbook(
        r'C:\Users\Administrator\Downloads\支付行为指数.xlsx')
    sheet1 = workbook.worksheets[0]
    count = 0
    print(sheet1.cell(1,4).value)
    text = '{'

    for row in range(5, 522):
        key = sheet1.cell(row, 6).value
        # print('key:', key)
        if (key):
            count = count + 1
            value = sheet1.cell(row, 7).value
            # print(key, ":", '"' + value + '",')
            s1 = '"{0}": "{1}",\n'.format(value, key)
            # s2 = '{0}: "{1}",\n'.format(value, value)
            text += s1       
    text += '}'
    with  open(r'C:\Users\Administrator\Desktop\test', 'w+', encoding='utf8') as f:
        f.write(text)
        
    print('count: ', count)


def generatorText():
    print('export const ListPlatformType = {')
    for line in TEXT.strip().splitlines():
        # prxint(line, '\n')
        a = line.split()
        k = a[0][3:]
        v = a[2].split('，')[1]
        print("{0}: '{1}',".format(k,v))

    print('}')
def generatorText1():
    print('export const ListType = {')
    for line in TEXT1.strip().splitlines():
        # print(line, '\n')
        a = line.split()
        k = a[0]
        v = a[2]
        print("{0}: '{1}',".format(k,v))

    print('}')
        
        


if __name__ == '__main__':
    # generatorText()
    read_excel()